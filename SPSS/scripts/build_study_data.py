from __future__ import annotations

import csv
import json
import os
import re
import shutil
import subprocess
import sys
from collections import OrderedDict
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


class ValidationError(RuntimeError):
    pass


ROOT = Path(__file__).resolve().parents[2]
SPSS_DIR = ROOT / "SPSS"
ANALYSIS_DIR = SPSS_DIR / "analysis"
QUANT_DIR = ANALYSIS_DIR / "quantitative"
QUAL_DIR = ANALYSIS_DIR / "qualitative"
DOC_DIR = ANALYSIS_DIR / "documentation"
TMP_DIR = ANALYSIS_DIR / "tmp"
FIGURES_DIR = ANALYSIS_DIR / "figures"

QUESTIONNAIRE_PATH = ROOT / "Latex" / "sections" / "appendix_questionnaire.tex"
FORMS_PATH = ROOT / "Documents" / "Methodology_and_Study" / "Private" / "vr_study_forms_latest.xlsx"
SESSION_NOTES_PATH = ROOT / "Documents" / "Methodology_and_Study" / "Private" / "participant_session_notes_private.md"
MASTER_GUIDE_PATH = ROOT / "Documents" / "Private_Support" / "VR_Study_Data_Analysis_Master_Guide.md"

SCRIPT_DIR = SPSS_DIR / "scripts"
NODE_EXECUTABLE = Path(
    __import__("os").environ.get(
        "CODEX_NODE_EXECUTABLE",
        r"C:\Users\Pri\.cache\codex-runtimes\codex-primary-runtime\dependencies\node\bin\node.exe",
    )
)
WORKBOOK_BUILDER_PATH = SCRIPT_DIR / "build_study_workbooks.mjs"
LIKERT_FIGURE_SCRIPT_PATH = SCRIPT_DIR / "generate_full_figure_set.py"
LIKERT_FIGURE_BASENAMES = (
    "Figure_01_Usability_Likert",
    "Figure_02_Engagement_Likert",
    "Figure_03_Learning_Experience_Likert",
    "Figure_04_Discomfort",
    "Figure_05_Post_Knowledge_Percent_Correct",
    "Figure_06_Baseline_vs_Post_Repeated_Knowledge",
)
LIKERT_PLOT_PYTHON = os.environ.get("CODEX_PLOT_PYTHON_EXECUTABLE") or shutil.which("python") or sys.executable


def normalize_text(value: Any) -> str:
    if value is None:
        return ""

    text = str(value)
    text = text.replace("\u00a0", " ").replace("\u200b", "")
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def normalize_key(value: Any) -> str:
    return normalize_text(value).casefold()


def quoted(text: str) -> str:
    return "'" + text.replace("'", "''") + "'"


def mean_if_complete(values: list[Any]) -> float | None:
    if any(value is None for value in values):
        return None

    return round(sum(values) / len(values), 4)


def sum_if_complete(values: list[Any]) -> int | None:
    if any(value is None for value in values):
        return None

    return int(sum(values))


def diff_if_complete(left: int | float | None, right: int | float | None) -> int | float | None:
    if left is None or right is None:
        return None

    result = right - left
    return round(result, 4) if isinstance(result, float) else result


def csv_scalar(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float):
        return f"{value:.4f}".rstrip("0").rstrip(".") if "." in f"{value:.4f}" else f"{value:.4f}"
    return str(value)


def split_semicolon_values(value: Any) -> list[str]:
    if value is None:
        return []

    return [normalize_text(part) for part in str(value).replace("\u00a0", " ").split(";") if normalize_text(part)]


@dataclass(frozen=True)
class McqQuestion:
    number: int
    prompt: str
    options: tuple[str, str, str, str]
    correct_code: int


MCQ_QUESTIONS: list[McqQuestion] = [
    McqQuestion(11, "Which register stores the address of the next instruction?", ("Stack Pointer", "Program Counter", "Instruction Register", "Base Register"), 2),
    McqQuestion(12, "Which stage of the instruction cycle reads an instruction from memory?", ("Decode", "Fetch", "Execute", "Write-back"), 2),
    McqQuestion(13, "During instruction execution, the instruction register primarily stores:", ("Program variables", "The currently executing instruction", "The stack pointer value", "Cache addresses"), 2),
    McqQuestion(14, "What happens to the program counter after an instruction is fetched?", ("It resets to zero", "It increments to the next instruction address", "It copies into the stack pointer", "It stores the result of the ALU"), 2),
    McqQuestion(15, "Which pipeline stage usually performs arithmetic operations?", ("Fetch", "Decode", "Execute", "Write-back"), 3),
    McqQuestion(16, "Branch instructions primarily affect which component?", ("Cache memory", "Program counter", "Stack pointer", "Instruction register"), 2),
    McqQuestion(17, "What does this instruction do? add $t0, $t1, $t2", ("Adds $t1 and $t2 and stores the result in $t0", "Adds $t0 and $t1 and stores the result in $t2", "Moves $t1 into $t2", "Multiplies registers"), 1),
    McqQuestion(18, "What does this instruction do? addi $t0, $zero, 10", ("Stores 10 into $t0", "Clears $t0", "Moves $t0 into $zero", "Adds two registers"), 1),
    McqQuestion(19, "Which instruction type loads data from memory into a register?", ("ADD", "STORE", "LOAD", "JUMP"), 3),
    McqQuestion(20, "What is the role of register $zero in MIPS?", ("Stores the program counter", "Always contains the value 0", "Stores return addresses", "Tracks the stack pointer"), 2),
    McqQuestion(21, "What does this instruction do? lw $t0, 0($s0)", ("Stores $t0 into memory", "Loads a word from memory into $t0", "Moves $s0 into $t0", "Clears memory"), 2),
    McqQuestion(22, "What does this instruction do? sw $t0, 4($s0)", ("Value from $t0 is stored in memory", "Value from memory is loaded into $t0", "$s0 is overwritten", "Stack pointer moves"), 1),
    McqQuestion(23, "What does this instruction do? beq $t0, $t1, label", ("Jump if $t0 > $t1", "Jump if $t0 == $t1", "Jump if $t0 < $t1", "Always jump"), 2),
    McqQuestion(24, "What does this instruction do? j label", ("Jumps to label unconditionally", "Jumps only if two registers are equal", "Stores the return address in $ra", "Loads a value from memory"), 1),
    McqQuestion(25, "Put the instruction stages in the correct order.", ("Execute, Fetch, Decode, Write-back", "Fetch, Decode, Execute, Write-back", "Decode, Fetch, Execute, Write-back", "Fetch, Execute, Decode, Write-back"), 2),
]

MCQ_BY_NUMBER = {question.number: question for question in MCQ_QUESTIONS}
ANSWER_KEY = {question.number: question.correct_code for question in MCQ_QUESTIONS}

FORMS_HEADER_MAP = OrderedDict(
    [
        ("ParticipantID", "id"),
        ("Email", "email"),
        ("Name", "name"),
        ("Q1", "are you 18 or older?"),
        ("Q2", "have you previously studied computer architecture or computer organisation?"),
        ("Q3", "have you used a vr headset before?"),
        ("Q4", "how experienced are you with virtual reality"),
        ("Q5", "before this session, how confident are you in your understanding of computer architecture concepts?"),
        ("Q6", "how difficult do you usually find topics such as instruction flow, pipelining, stack behaviour, and memory hierarchy?"),
        ("Q7", "which of the following topics have you previously studied?"),
        ("USE01", "the vr system was easy to learn"),
        ("USE02", "i could navigate the environment without difficulty"),
        ("USE03", "the controls felt intuitive"),
        ("USE04", "the visual presentation made the concepts easier to follow"),
        ("USE05", "i could focus on the learning task without being distracted by the interface"),
        ("USE06", "i would feel confident using this system again"),
        ("USE07_Discomfort", "i experienced discomfort, motion sickness, or visual fatigue while using the system"),
        ("ENG01", "the experience held my attention"),
        ("ENG02", "i felt actively involved in the learning process"),
        ("ENG03", "the system made the topic feel more interesting"),
        ("ENG04", "i was motivated to explore the concepts further"),
        ("ENG05", "the interactive format improved my focus compared with standard learning materials"),
        ("LEARN01", "the vr experience helped me understand the topic better"),
        ("LEARN02", "the visualisations helped me build a mental model of what was happening"),
        ("LEARN03", "the system made abstract processes easier to understand"),
        ("LEARN04", "i would prefer this as a supplement to traditional teaching materials"),
        ("LEARN05", "i believe i learned something useful from this session"),
        ("POST_Q11", "which register stores the address of the next instruction?"),
        ("POST_Q12", "which stage of the instruction cycle reads an instruction from memory?"),
        ("POST_Q13", "during instruction execution, the instruction register primarily stores:"),
        ("POST_Q14", "what happens to the program counter after an instruction is fetched?"),
        ("POST_Q15", "which pipeline stage usually performs arithmetic operations?"),
        ("POST_Q16", "branch instructions primarily affect which component?"),
        ("POST_Q17", "what does this instruction do? add $t0, $t1, $t2"),
        ("POST_Q18", "what does this instruction do? addi $t0, $zero, 10"),
        ("POST_Q19", "which instruction type loads data from memory into a register?"),
        ("POST_Q20", "what is the role of register $zero in mips?"),
        ("POST_Q21", "what does this instruction do? lw $t0, 0($s0)"),
        ("POST_Q22", "what does this instruction do? sw $t0, 4($s0)"),
        ("POST_Q23", "what does this instruction do? beq $t0, $t1, label"),
        ("POST_Q24", "what does this instruction do? j label"),
        ("POST_Q25", "put the instruction stages in the correct order."),
        ("POST_Q26", "how confident are you in your answers to the previous section?"),
        ("Q27", "what part of the experience helped your understanding most?"),
        ("Q28", "what part was confusing or ineffective?"),
        ("Q29", "what would you improve in the system?"),
        ("Q30", "were there any moments where the vr interface got in the way of learning?"),
        ("Q31", "would you use this again as a learning aid? why or why not?"),
    ]
)


QUESTIONNAIRE_CHECKS = [
    r"\\question{1\. Are you 18 or older\?}",
    r"\\question{8\. Please indicate how strongly you agree with the following statements for System Usability\.}",
    r"\\question{10\. Please indicate how strongly you agree with the following statements about your learning experience\.}",
    r"\\question{25\. Put the instruction stages in the correct order\.}",
    r"\\question{31\. Would you use this again as a learning aid\? Why or why not\?}",
]


YES_NO = {"yes": 1, "no": 0}
VR_EXPERIENCE = {"none": 1, "limited": 2, "moderate": 3, "good": 4, "high": 5}
PRIOR_CONFIDENCE = {
    "extremely not confident": 1,
    "somewhat not confident": 2,
    "neutral": 3,
    "somewhat confident": 4,
    "extremely confident": 5,
}
PRIOR_DIFFICULTY = {
    "extremely difficult": 1,
    "somewhat difficult": 2,
    "neutral": 3,
    "somewhat easy": 4,
    "extremely easy": 5,
}
LIKERT = {"strongly disagree": 1, "disagree": 2, "neutral": 3, "agree": 4, "strongly agree": 5}
KNOWLEDGE_CONFIDENCE = {
    "1": 1,
    "2": 2,
    "3": 3,
    "4": 4,
    "5": 5,
    "not at all confident": 1,
    "very confident": 5,
}

TOPIC_VARIABLES = OrderedDict(
    [
        ("Instruction Flow", "BG_Studied_InstructionFlow"),
        ("Pipelining", "BG_Studied_Pipelining"),
        ("Stack", "BG_Studied_Stack"),
        ("Memory Hierarchy", "BG_Studied_MemoryHierarchy"),
        ("Assembly", "BG_Studied_Assembly"),
        ("None of the above", "BG_Studied_None"),
    ]
)

OPEN_ENDED_PROMPTS = OrderedDict(
    [
        ("Q27", "What part of the experience helped your understanding most?"),
        ("Q28", "What part was confusing or ineffective?"),
        ("Q29", "What would you improve in the system?"),
        ("Q30", "Were there any moments where the VR interface got in the way of learning?"),
        ("Q31", "Would you use this again as a learning aid? Why or why not?"),
    ]
)

BASELINE_ALIAS_MAP = {
    normalize_key("PC"): "Program Counter",
    normalize_key("SP"): "Stack Pointer",
    normalize_key("Current Instruction"): "The currently executing instruction",
    normalize_key("current"): "The currently executing instruction",
    normalize_key("curr instruction"): "The currently executing instruction",
    normalize_key("Program Variables"): "Program variables",
    normalize_key("program variable"): "Program variables",
    normalize_key("increments"): "It increments to the next instruction address",
    normalize_key("reset to 0"): "It resets to zero",
    normalize_key("resets to zero"): "It resets to zero",
}


BASELINE_FIELD_TO_QUESTION = OrderedDict(
    [
        ("PRE_Q11_RESP", 11),
        ("PRE_Q12_RESP", 12),
        ("PRE_Q13_RESP", 13),
        ("PRE_Q14_RESP", 14),
        ("PRE_Q25_RESP", 25),
    ]
)

BASELINE_QUESTIONS = tuple(BASELINE_FIELD_TO_QUESTION.values())
MATCHED_KNOWLEDGE5_QUESTIONS = (11, 12, 13, 14, 25)


def mcq_normalized_variants(question: McqQuestion) -> dict[str, int]:
    variants: dict[str, int] = {}

    for code, option in enumerate(question.options, start=1):
        normalized = normalize_key(option)
        variants[normalized] = code

        if question.number == 25:
            pieces = [normalize_text(piece) for piece in option.split(",")]
            compact = ";".join(pieces)
            variants[normalize_key(compact)] = code
            variants[normalize_key(compact + ";")] = code
            variants[normalize_key("; ".join(pieces))] = code
            variants[normalize_key("; ".join(pieces) + ";")] = code

    return variants


MCQ_RESPONSE_VARIANTS = {question.number: mcq_normalized_variants(question) for question in MCQ_QUESTIONS}


VARIABLE_ORDER = [
    "ParticipantID",
    "BG_Age18Plus",
    "BG_PriorArchitecture",
    "BG_PriorVR",
    "BG_VRExperience",
    "BG_PriorConfidence",
    "BG_PriorDifficulty",
    "BG_Studied_InstructionFlow",
    "BG_Studied_Pipelining",
    "BG_Studied_Stack",
    "BG_Studied_MemoryHierarchy",
    "BG_Studied_Assembly",
    "BG_Studied_None",
    "USE01",
    "USE02",
    "USE03",
    "USE04",
    "USE05",
    "USE06",
    "USE07_Discomfort",
    "USE07_Comfort_R",
    "UsabilityMean_6",
    "UsabilityMean_7R",
    "ENG01",
    "ENG02",
    "ENG03",
    "ENG04",
    "ENG05",
    "EngagementMean",
    "LEARN01",
    "LEARN02",
    "LEARN03",
    "LEARN04",
    "LEARN05",
    "LearningExperienceMean",
    "PRE_Q11_RESP",
    "PRE_Q12_RESP",
    "PRE_Q13_RESP",
    "PRE_Q14_RESP",
    "PRE_Q25_RESP",
    "PRE_Q11_CORR",
    "PRE_Q12_CORR",
    "PRE_Q13_CORR",
    "PRE_Q14_CORR",
    "PRE_Q25_CORR",
    "POST_Q11_RESP",
    "POST_Q12_RESP",
    "POST_Q13_RESP",
    "POST_Q14_RESP",
    "POST_Q15_RESP",
    "POST_Q16_RESP",
    "POST_Q17_RESP",
    "POST_Q18_RESP",
    "POST_Q19_RESP",
    "POST_Q20_RESP",
    "POST_Q21_RESP",
    "POST_Q22_RESP",
    "POST_Q23_RESP",
    "POST_Q24_RESP",
    "POST_Q25_RESP",
    "POST_Q11_CORR",
    "POST_Q12_CORR",
    "POST_Q13_CORR",
    "POST_Q14_CORR",
    "POST_Q15_CORR",
    "POST_Q16_CORR",
    "POST_Q17_CORR",
    "POST_Q18_CORR",
    "POST_Q19_CORR",
    "POST_Q20_CORR",
    "POST_Q21_CORR",
    "POST_Q22_CORR",
    "POST_Q23_CORR",
    "POST_Q24_CORR",
    "POST_Q25_CORR",
    "PRE_Knowledge5_NValid",
    "PRE_Knowledge5_Total",
    "POST_Knowledge5_Total",
    "Knowledge5_Gain",
    "KnowledgeMatched_Available_NValid",
    "PRE_KnowledgeMatched_Available_Total",
    "POST_KnowledgeMatched_Available_Total",
    "KnowledgeMatched_Available_Gain",
    "PRE_KnowledgeMatched_Available_Percent",
    "POST_KnowledgeMatched_Available_Percent",
    "KnowledgeMatched_Available_GainPercent",
    "POST_Knowledge15_Total",
    "POST_Knowledge15_Percent",
    "POST_KnowledgeConfidence",
]


def verify_source_files() -> list[str]:
    messages: list[str] = []

    questionnaire_text = QUESTIONNAIRE_PATH.read_text(encoding="utf-8")
    for pattern in QUESTIONNAIRE_CHECKS:
        if not re.search(pattern, questionnaire_text):
            raise ValidationError(f"Questionnaire verification failed for pattern: {pattern}")
    messages.append("appendix_questionnaire.tex: verified representative question blocks and wording anchors.")

    if not FORMS_PATH.exists():
        raise ValidationError(f"Forms export not found: {FORMS_PATH}")
    if not SESSION_NOTES_PATH.exists():
        raise ValidationError(f"Session notes not found: {SESSION_NOTES_PATH}")
    if not MASTER_GUIDE_PATH.exists():
        raise ValidationError(f"Master guide not found: {MASTER_GUIDE_PATH}")

    workbook = load_workbook(FORMS_PATH, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    headers = [normalize_key(cell.value) for cell in worksheet[1]]
    missing_headers = [key for key, expected in FORMS_HEADER_MAP.items() if expected not in headers]
    if missing_headers:
        raise ValidationError(f"Forms workbook is missing expected columns: {', '.join(missing_headers)}")
    messages.append(f"{FORMS_PATH.name}: verified workbook structure with {worksheet.max_row - 1} response rows and {worksheet.max_column} columns.")

    notes_text = SESSION_NOTES_PATH.read_text(encoding="utf-8")
    participant_headers = re.findall(r"^## Participant \d+\s*$", notes_text, flags=re.MULTILINE)
    if not participant_headers:
        raise ValidationError("Session notes do not contain any '## Participant N' sections.")
    messages.append(f"{SESSION_NOTES_PATH.name}: verified {len(participant_headers)} participant sections.")

    return messages


def load_forms_rows() -> tuple[list[dict[str, Any]], list[str]]:
    workbook = load_workbook(FORMS_PATH, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    header_indices = {normalize_key(cell.value): index for index, cell in enumerate(worksheet[1], start=1)}

    forms_rows: list[dict[str, Any]] = []
    seen_ids: set[int] = set()
    warnings: list[str] = []

    for row_number in range(2, worksheet.max_row + 1):
        raw_id = worksheet.cell(row=row_number, column=header_indices[FORMS_HEADER_MAP["ParticipantID"]]).value
        if raw_id in (None, ""):
            continue

        try:
            participant_id = int(raw_id)
        except (TypeError, ValueError) as exc:
            raise ValidationError(f"ParticipantID at workbook row {row_number} is not an integer: {raw_id!r}") from exc

        if participant_id in seen_ids:
            raise ValidationError(f"Duplicate ParticipantID in Forms workbook: {participant_id}")
        seen_ids.add(participant_id)

        extracted = {"ParticipantID": participant_id}
        for target_key, header_key in FORMS_HEADER_MAP.items():
            column_index = header_indices[header_key]
            extracted[target_key] = worksheet.cell(row=row_number, column=column_index).value

        if participant_id == 8 and not normalize_text(extracted["Email"]) and not normalize_text(extracted["Name"]):
            warnings.append("Participant 8 is anonymous in the Forms export; direct identifiers were blank in the source workbook.")

        forms_rows.append(extracted)

    if not forms_rows:
        raise ValidationError("No Forms participants were loaded from the workbook.")

    return forms_rows, warnings


def parse_session_notes() -> dict[int, dict[str, Any]]:
    text = SESSION_NOTES_PATH.read_text(encoding="utf-8")
    lines = text.splitlines()
    sections: dict[int, dict[str, Any]] = {}
    current_id: int | None = None
    current_field: str | None = None

    participant_re = re.compile(r"^## Participant (\d+)\s*$")
    field_re = re.compile(r"^- ([^:]+):\s*(.*)$")
    nested_re = re.compile(r"^\s{2}-\s*(.*)$")

    for line in lines:
        participant_match = participant_re.match(line)
        if participant_match:
            current_id = int(participant_match.group(1))
            if current_id in sections:
                raise ValidationError(f"Duplicate participant section in session notes: {current_id}")
            sections[current_id] = {}
            current_field = None
            continue

        if current_id is None:
            continue

        field_match = field_re.match(line)
        if field_match:
            current_field = field_match.group(1).strip()
            value = field_match.group(2).strip()
            sections[current_id][current_field] = value if value else []
            continue

        nested_match = nested_re.match(line)
        if nested_match and current_field is not None:
            entry = nested_match.group(1).strip()
            existing = sections[current_id].get(current_field, [])
            if not isinstance(existing, list):
                existing = [existing] if normalize_text(existing) else []
            existing.append(entry)
            sections[current_id][current_field] = existing

    if not sections:
        raise ValidationError("Session notes parsing failed; no participant sections were captured.")

    return sections


def encode_lookup(raw_value: Any, mapping: dict[str, int], field_name: str) -> int | None:
    normalized = normalize_key(raw_value)
    if not normalized:
        return None
    if normalized not in mapping:
        raise ValidationError(f"Unknown value for {field_name}: {raw_value!r}")
    return mapping[normalized]


def encode_topics(raw_value: Any) -> dict[str, int]:
    selected = split_semicolon_values(raw_value)
    selected_keys = {normalize_key(item) for item in selected}
    canonical_map = {normalize_key(source): variable for source, variable in TOPIC_VARIABLES.items()}

    invalid = sorted(item for item in selected_keys if item not in canonical_map)
    if invalid:
        raise ValidationError(f"Unknown Q7 topic selections: {invalid}")

    output = {variable: 0 for variable in TOPIC_VARIABLES.values()}
    for source, variable in TOPIC_VARIABLES.items():
        if normalize_key(source) in selected_keys:
            output[variable] = 1

    if output["BG_Studied_None"] == 1 and sum(output[var] for var in output if var != "BG_Studied_None") > 0:
        raise ValidationError("Q7 contains 'None of the above' alongside one or more real topics.")

    return output


def map_post_mcq_response(question_number: int, raw_value: Any) -> int | None:
    normalized = normalize_key(raw_value)
    if not normalized:
        return None

    variants = MCQ_RESPONSE_VARIANTS[question_number]
    if normalized not in variants:
        raise ValidationError(f"MCQ response for Q{question_number} could not be mapped: {raw_value!r}")
    return variants[normalized]


def map_baseline_response(question_number: int, raw_value: str) -> int | None:
    normalized = normalize_key(raw_value)
    if not normalized:
        return None
    if normalized == "n/a":
        return None
    if normalized in {"1", "2", "3", "4"}:
        return int(normalized)

    canonical = BASELINE_ALIAS_MAP.get(normalized)
    if canonical is None:
        options = {normalize_key(option): code for code, option in enumerate(MCQ_BY_NUMBER[question_number].options, start=1)}
        if normalized not in options:
            raise ValidationError(f"Baseline response for Q{question_number} could not be mapped: {raw_value!r}")
        return options[normalized]

    options = {normalize_key(option): code for code, option in enumerate(MCQ_BY_NUMBER[question_number].options, start=1)}
    normalized_canonical = normalize_key(canonical)
    if normalized_canonical not in options:
        raise ValidationError(f"Canonical baseline mapping failed for Q{question_number}: {raw_value!r} -> {canonical!r}")
    return options[normalized_canonical]


def build_quantitative_rows(forms_rows: list[dict[str, Any]], baseline_sections: dict[int, dict[str, Any]]) -> tuple[list[OrderedDict[str, Any]], list[dict[str, Any]], list[dict[str, Any]], dict[str, int], list[str]]:
    baseline_ids = set(baseline_sections)
    forms_ids = {int(row["ParticipantID"]) for row in forms_rows}

    if forms_ids != baseline_ids:
        raise ValidationError(
            "Forms and baseline participant IDs do not align as expected. "
            f"Forms IDs: {sorted(forms_ids)} | Baseline IDs: {sorted(baseline_ids)}"
        )

    quantitative_rows: list[OrderedDict[str, Any]] = []
    questionnaire_feedback_rows: list[dict[str, Any]] = []
    session_note_rows: list[dict[str, Any]] = []
    baseline_missing_counts: dict[str, int] = {
        "PRE_Q11_RESP": 0,
        "PRE_Q12_RESP": 0,
        "PRE_Q13_RESP": 0,
        "PRE_Q14_RESP": 0,
        "PRE_Q25_RESP": 0,
    }
    baseline_warnings: list[str] = []

    for forms_row in sorted(forms_rows, key=lambda item: int(item["ParticipantID"])):
        participant_id = int(forms_row["ParticipantID"])
        notes = baseline_sections[participant_id]
        encoded: OrderedDict[str, Any] = OrderedDict((variable, None) for variable in VARIABLE_ORDER)
        encoded["ParticipantID"] = participant_id

        encoded["BG_Age18Plus"] = encode_lookup(forms_row["Q1"], YES_NO, "Q1")
        encoded["BG_PriorArchitecture"] = encode_lookup(forms_row["Q2"], YES_NO, "Q2")
        encoded["BG_PriorVR"] = encode_lookup(forms_row["Q3"], YES_NO, "Q3")
        encoded["BG_VRExperience"] = encode_lookup(forms_row["Q4"], VR_EXPERIENCE, "Q4")
        encoded["BG_PriorConfidence"] = encode_lookup(forms_row["Q5"], PRIOR_CONFIDENCE, "Q5")
        encoded["BG_PriorDifficulty"] = encode_lookup(forms_row["Q6"], PRIOR_DIFFICULTY, "Q6")
        encoded.update(encode_topics(forms_row["Q7"]))

        for key in ("USE01", "USE02", "USE03", "USE04", "USE05", "USE06", "USE07_Discomfort", "ENG01", "ENG02", "ENG03", "ENG04", "ENG05", "LEARN01", "LEARN02", "LEARN03", "LEARN04", "LEARN05"):
            encoded[key] = encode_lookup(forms_row[key], LIKERT, key)

        encoded["USE07_Comfort_R"] = 6 - encoded["USE07_Discomfort"] if encoded["USE07_Discomfort"] is not None else None
        encoded["UsabilityMean_6"] = mean_if_complete([encoded[f"USE0{i}"] for i in range(1, 7)])
        encoded["UsabilityMean_7R"] = mean_if_complete([encoded[f"USE0{i}"] for i in range(1, 7)] + [encoded["USE07_Comfort_R"]])
        encoded["EngagementMean"] = mean_if_complete([encoded[f"ENG0{i}"] for i in range(1, 6)])
        encoded["LearningExperienceMean"] = mean_if_complete([encoded[f"LEARN0{i}"] for i in range(1, 6)])

        prescreen_values_raw = notes.get("PreScreen Answers")
        if not prescreen_values_raw or isinstance(prescreen_values_raw, list):
            raise ValidationError(f"Participant {participant_id} is missing a parseable 'PreScreen Answers' field.")
        prescreen_values = [part.strip() for part in str(prescreen_values_raw).split(",")]
        if len(prescreen_values) != 5:
            raise ValidationError(f"Participant {participant_id} has malformed PreScreen Answers: {prescreen_values_raw!r}")

        for pre_var, question_number in BASELINE_FIELD_TO_QUESTION.items():
            mapped = map_baseline_response(question_number, prescreen_values[list(BASELINE_FIELD_TO_QUESTION).index(pre_var)])
            encoded[pre_var] = mapped
            if mapped is None:
                baseline_missing_counts[pre_var] += 1
            corr_var = pre_var.replace("_RESP", "_CORR")
            encoded[corr_var] = None if mapped is None else int(mapped == ANSWER_KEY[question_number])

        for question in MCQ_QUESTIONS:
            raw_response = forms_row[f"POST_Q{question.number}"]
            resp_var = f"POST_Q{question.number}_RESP"
            corr_var = f"POST_Q{question.number}_CORR"
            encoded[resp_var] = map_post_mcq_response(question.number, raw_response)
            encoded[corr_var] = None if encoded[resp_var] is None else int(encoded[resp_var] == question.correct_code)

        encoded["POST_KnowledgeConfidence"] = encode_lookup(forms_row["POST_Q26"], KNOWLEDGE_CONFIDENCE, "POST_Q26")

        encoded["PRE_Knowledge5_NValid"] = sum(1 for number in MATCHED_KNOWLEDGE5_QUESTIONS if encoded[f"PRE_Q{number}_CORR"] is not None)
        encoded["PRE_Knowledge5_Total"] = sum_if_complete([encoded[f"PRE_Q{number}_CORR"] for number in MATCHED_KNOWLEDGE5_QUESTIONS])
        encoded["POST_Knowledge5_Total"] = sum_if_complete([encoded[f"POST_Q{number}_CORR"] for number in MATCHED_KNOWLEDGE5_QUESTIONS])
        encoded["Knowledge5_Gain"] = diff_if_complete(encoded["PRE_Knowledge5_Total"], encoded["POST_Knowledge5_Total"])
        available_matched_questions = [number for number in MATCHED_KNOWLEDGE5_QUESTIONS if encoded[f"PRE_Q{number}_CORR"] is not None]
        encoded["KnowledgeMatched_Available_NValid"] = len(available_matched_questions)
        encoded["PRE_KnowledgeMatched_Available_Total"] = (
            sum(encoded[f"PRE_Q{number}_CORR"] for number in available_matched_questions)
            if available_matched_questions
            else None
        )
        encoded["POST_KnowledgeMatched_Available_Total"] = (
            sum_if_complete([encoded[f"POST_Q{number}_CORR"] for number in available_matched_questions])
            if available_matched_questions
            else None
        )
        encoded["KnowledgeMatched_Available_Gain"] = diff_if_complete(
            encoded["PRE_KnowledgeMatched_Available_Total"],
            encoded["POST_KnowledgeMatched_Available_Total"],
        )
        encoded["PRE_KnowledgeMatched_Available_Percent"] = (
            round(encoded["PRE_KnowledgeMatched_Available_Total"] / encoded["KnowledgeMatched_Available_NValid"] * 100, 4)
            if encoded["PRE_KnowledgeMatched_Available_Total"] is not None and encoded["KnowledgeMatched_Available_NValid"] > 0
            else None
        )
        encoded["POST_KnowledgeMatched_Available_Percent"] = (
            round(encoded["POST_KnowledgeMatched_Available_Total"] / encoded["KnowledgeMatched_Available_NValid"] * 100, 4)
            if encoded["POST_KnowledgeMatched_Available_Total"] is not None and encoded["KnowledgeMatched_Available_NValid"] > 0
            else None
        )
        encoded["KnowledgeMatched_Available_GainPercent"] = diff_if_complete(
            encoded["PRE_KnowledgeMatched_Available_Percent"],
            encoded["POST_KnowledgeMatched_Available_Percent"],
        )
        encoded["POST_Knowledge15_Total"] = sum_if_complete([encoded[f"POST_Q{number}_CORR"] for number in range(11, 26)])
        encoded["POST_Knowledge15_Percent"] = round(encoded["POST_Knowledge15_Total"] / 15 * 100, 4) if encoded["POST_Knowledge15_Total"] is not None else None

        quantitative_rows.append(encoded)

        for question_id, prompt in OPEN_ENDED_PROMPTS.items():
            questionnaire_feedback_rows.append(
                {
                    "ParticipantID": participant_id,
                    "QuestionID": question_id,
                    "Prompt": prompt,
                    "Response": normalize_text(forms_row[question_id]),
                    "SourceType": "ParticipantFeedback",
                }
            )

        append_session_notes_rows(session_note_rows, participant_id, notes)

    baseline_warnings.append(
        "Baseline PreScreen Answers are interpreted strictly as Q11, Q12, Q13, Q14, Q25. No baseline confidence variable is created because Q26 confidence is only collected after the VR session."
    )

    return quantitative_rows, questionnaire_feedback_rows, session_note_rows, baseline_missing_counts, baseline_warnings


def append_session_notes_rows(output_rows: list[dict[str, Any]], participant_id: int, notes: dict[str, Any]) -> None:
    field_config = [
        ("Instructions / modes attempted", "SessionMetadata", "ModesAttempted"),
        ("Instruction / modes attempted", "SessionMetadata", "ModesAttempted"),
        ("Background / refresher", "SessionMetadata", "BackgroundRefresher"),
        ("Session observations", "ResearcherObservation", "SessionObservation"),
        ("Useful feedback to retain", "ResearcherObservation", "UsefulFeedback"),
        ("Notes to clean up later", "ResearcherObservation", "CleanupNote"),
    ]

    for field_name, source_type, note_category in field_config:
        if field_name not in notes:
            continue

        field_value = notes[field_name]
        items = field_value if isinstance(field_value, list) else [field_value]
        for item in items:
            text = normalize_text(item)
            if not text:
                continue
            output_rows.append(
                {
                    "ParticipantID": participant_id,
                    "SourceType": source_type,
                    "NoteCategory": note_category,
                    "Text": text,
                }
            )


def validate_quantitative_rows(rows: list[OrderedDict[str, Any]]) -> dict[str, dict[str, int]]:
    if not rows:
        raise ValidationError("No quantitative rows were built.")

    ids = [row["ParticipantID"] for row in rows]
    if len(ids) != len(set(ids)):
        raise ValidationError("Duplicate ParticipantID values appeared in the quantitative dataset.")

    binary_fields = [
        "BG_Age18Plus",
        "BG_PriorArchitecture",
        "BG_PriorVR",
        "BG_Studied_InstructionFlow",
        "BG_Studied_Pipelining",
        "BG_Studied_Stack",
        "BG_Studied_MemoryHierarchy",
        "BG_Studied_Assembly",
        "BG_Studied_None",
    ]

    ordinal_1_to_5 = [
        "BG_VRExperience",
        "BG_PriorConfidence",
        "BG_PriorDifficulty",
        "USE01",
        "USE02",
        "USE03",
        "USE04",
        "USE05",
        "USE06",
        "USE07_Discomfort",
        "ENG01",
        "ENG02",
        "ENG03",
        "ENG04",
        "ENG05",
        "LEARN01",
        "LEARN02",
        "LEARN03",
        "LEARN04",
        "LEARN05",
        "POST_KnowledgeConfidence",
    ]

    for row in rows:
        for field in binary_fields:
            if row[field] not in (0, 1, None):
                raise ValidationError(f"{field} contains an invalid value for ParticipantID {row['ParticipantID']}: {row[field]!r}")
        for field in ordinal_1_to_5:
            if row[field] is not None and not (1 <= row[field] <= 5):
                raise ValidationError(f"{field} is outside 1-5 for ParticipantID {row['ParticipantID']}: {row[field]!r}")

        for question_number in BASELINE_QUESTIONS:
            resp = row[f"PRE_Q{question_number}_RESP"]
            corr = row[f"PRE_Q{question_number}_CORR"]
            if resp is not None and not (1 <= resp <= 4):
                raise ValidationError(f"PRE_Q{question_number}_RESP is outside 1-4 for ParticipantID {row['ParticipantID']}: {resp!r}")
            if corr not in (0, 1, None):
                raise ValidationError(f"PRE_Q{question_number}_CORR contains an invalid value for ParticipantID {row['ParticipantID']}: {corr!r}")

        for question_number in range(11, 26):
            resp = row[f"POST_Q{question_number}_RESP"]
            corr = row[f"POST_Q{question_number}_CORR"]
            if resp is not None and not (1 <= resp <= 4):
                raise ValidationError(f"POST_Q{question_number}_RESP is outside 1-4 for ParticipantID {row['ParticipantID']}: {resp!r}")
            if corr not in (0, 1, None):
                raise ValidationError(f"POST_Q{question_number}_CORR contains an invalid value for ParticipantID {row['ParticipantID']}: {corr!r}")

        derived_ranges = {
            "USE07_Comfort_R": (1, 5),
            "UsabilityMean_6": (1, 5),
            "UsabilityMean_7R": (1, 5),
            "EngagementMean": (1, 5),
            "LearningExperienceMean": (1, 5),
            "PRE_Knowledge5_NValid": (0, 5),
            "PRE_Knowledge5_Total": (0, 5),
            "POST_Knowledge5_Total": (0, 5),
            "Knowledge5_Gain": (-5, 5),
            "KnowledgeMatched_Available_NValid": (0, 5),
            "PRE_KnowledgeMatched_Available_Total": (0, 5),
            "POST_KnowledgeMatched_Available_Total": (0, 5),
            "KnowledgeMatched_Available_Gain": (-5, 5),
            "PRE_KnowledgeMatched_Available_Percent": (0, 100),
            "POST_KnowledgeMatched_Available_Percent": (0, 100),
            "KnowledgeMatched_Available_GainPercent": (-100, 100),
            "POST_Knowledge15_Total": (0, 15),
            "POST_Knowledge15_Percent": (0, 100),
        }
        for field, (minimum, maximum) in derived_ranges.items():
            value = row[field]
            if value is not None and not (minimum <= value <= maximum):
                raise ValidationError(f"{field} is outside {minimum}-{maximum} for ParticipantID {row['ParticipantID']}: {value!r}")

    audit: dict[str, dict[str, int]] = {}
    for field in VARIABLE_ORDER:
        values = [row[field] for row in rows]
        audit[field] = {
            "valid": sum(value is not None for value in values),
            "missing": sum(value is None for value in values),
            "unexpected": 0,
        }

    return audit


def build_variable_metadata() -> list[dict[str, str]]:
    metadata: list[dict[str, str]] = []

    def add(variable: str, source: str, label: str, var_type: str, valid_values: str, level: str, derivation: str = "") -> None:
        metadata.append(
            {
                "Variable": variable,
                "Source": source,
                "Label": label,
                "Type": var_type,
                "ValidValues": valid_values,
                "Level": level,
                "Derivation": derivation,
            }
        )

    def mcq_option_codes(question: McqQuestion, allow_missing: bool = False) -> str:
        option_map = "; ".join(f"{index}={option}" for index, option in enumerate(question.options, start=1))
        return f"{option_map}; blank if missing" if allow_missing else option_map

    add("ParticipantID", "Forms Q0", "Participant ID", "Numeric", "Positive integer", "Nominal")
    add("BG_Age18Plus", "Q1", "Age 18 or older", "Numeric", "0=No; 1=Yes", "Nominal")
    add("BG_PriorArchitecture", "Q2", "Previously studied computer architecture or computer organisation", "Numeric", "0=No; 1=Yes", "Nominal")
    add("BG_PriorVR", "Q3", "Used a VR headset before", "Numeric", "0=No; 1=Yes", "Nominal")
    add("BG_VRExperience", "Q4", "VR experience level", "Numeric", "1=None; 2=Limited; 3=Moderate; 4=Good; 5=High", "Ordinal")
    add("BG_PriorConfidence", "Q5", "Pre-session confidence in computer architecture", "Numeric", "1=Extremely not confident ... 5=Extremely confident", "Ordinal")
    add("BG_PriorDifficulty", "Q6", "Perceived difficulty of core computer architecture topics", "Numeric", "1=Extremely difficult ... 5=Extremely easy", "Ordinal")
    add("BG_Studied_InstructionFlow", "Q7", "Previously studied: Instruction Flow", "Numeric", "0=Not selected; 1=Selected", "Nominal")
    add("BG_Studied_Pipelining", "Q7", "Previously studied: Pipelining", "Numeric", "0=Not selected; 1=Selected", "Nominal")
    add("BG_Studied_Stack", "Q7", "Previously studied: Stack", "Numeric", "0=Not selected; 1=Selected", "Nominal")
    add("BG_Studied_MemoryHierarchy", "Q7", "Previously studied: Memory Hierarchy", "Numeric", "0=Not selected; 1=Selected", "Nominal")
    add("BG_Studied_Assembly", "Q7", "Previously studied: Assembly", "Numeric", "0=Not selected; 1=Selected", "Nominal")
    add("BG_Studied_None", "Q7", "Previously studied: None of the above", "Numeric", "0=Not selected; 1=Selected", "Nominal")

    for field, label in [
        ("USE01", "The VR system was easy to learn"),
        ("USE02", "I could navigate the environment without difficulty"),
        ("USE03", "The controls felt intuitive"),
        ("USE04", "The visual presentation made the concepts easier to follow"),
        ("USE05", "I could focus on the learning task without being distracted by the interface"),
        ("USE06", "I would feel confident using this system again"),
        ("USE07_Discomfort", "I experienced discomfort, motion sickness, or visual fatigue while using the system"),
        ("ENG01", "The experience held my attention"),
        ("ENG02", "I felt actively involved in the learning process"),
        ("ENG03", "The system made the topic feel more interesting"),
        ("ENG04", "I was motivated to explore the concepts further"),
        ("ENG05", "The interactive format improved my focus compared with standard learning materials"),
        ("LEARN01", "The VR experience helped me understand the topic better"),
        ("LEARN02", "The visualisations helped me build a mental model of what was happening"),
        ("LEARN03", "The system made abstract processes easier to understand"),
        ("LEARN04", "I would prefer this as a supplement to traditional teaching materials"),
        ("LEARN05", "I believe I learned something useful from this session"),
    ]:
        add(field, field.replace("_", " ").split("0")[0], label, "Numeric", "1=Strongly Disagree ... 5=Strongly Agree", "Ordinal")

    add("USE07_Comfort_R", "Derived from USE07_Discomfort", "Reverse-coded comfort item", "Numeric", "1=Very uncomfortable ... 5=Very comfortable", "Ordinal", "6 - USE07_Discomfort")
    add("UsabilityMean_6", "Derived from USE01-USE06", "Mean of six usability items", "Numeric", "1-5", "Scale", "Mean(USE01:USE06) when complete")
    add("UsabilityMean_7R", "Derived from USE01-USE06 + USE07_Comfort_R", "Mean of six usability items plus reverse-coded comfort", "Numeric", "1-5", "Scale", "Mean(USE01:USE06, USE07_Comfort_R) when complete")
    add("EngagementMean", "Derived from ENG01-ENG05", "Mean of five engagement items", "Numeric", "1-5", "Scale", "Mean(ENG01:ENG05) when complete")
    add("LearningExperienceMean", "Derived from LEARN01-LEARN05", "Mean of five learning-experience items", "Numeric", "1-5", "Scale", "Mean(LEARN01:LEARN05) when complete")

    for question_number in BASELINE_QUESTIONS:
        question = MCQ_BY_NUMBER[question_number]
        add(
            f"PRE_Q{question_number}_RESP",
            "Private baseline notes",
            f"Baseline response for Q{question_number}: {question.prompt}",
            "Numeric",
            mcq_option_codes(question, allow_missing=True),
            "Nominal",
        )
    for question_number in BASELINE_QUESTIONS:
        add(f"PRE_Q{question_number}_CORR", "Derived from PRE_Q response", f"Baseline correctness for Q{question_number}", "Numeric", "0=Incorrect; 1=Correct; blank if missing", "Nominal")
    for question in MCQ_QUESTIONS:
        add(
            f"POST_Q{question.number}_RESP",
            f"Forms Q{question.number}",
            f"Post-session response for Q{question.number}: {question.prompt}",
            "Numeric",
            mcq_option_codes(question),
            "Nominal",
        )
    for question in MCQ_QUESTIONS:
        add(f"POST_Q{question.number}_CORR", f"Derived from POST_Q{question.number}_RESP", f"Post-session correctness for Q{question.number}", "Numeric", "0=Incorrect; 1=Correct", "Nominal")

    add("PRE_Knowledge5_NValid", "Derived from PRE_Q11_CORR, PRE_Q12_CORR, PRE_Q13_CORR, PRE_Q14_CORR, PRE_Q25_CORR", "Count of nonmissing baseline items among Q11, Q12, Q13, Q14, and Q25", "Numeric", "0-5", "Scale", "Count nonmissing PRE_Q11_CORR, PRE_Q12_CORR, PRE_Q13_CORR, PRE_Q14_CORR, PRE_Q25_CORR")
    add("PRE_Knowledge5_Total", "Derived from matched baseline responses", "Baseline five-item matched knowledge total", "Numeric", "0-5; blank if incomplete", "Scale", "Sum of PRE_Q11_CORR, PRE_Q12_CORR, PRE_Q13_CORR, PRE_Q14_CORR, PRE_Q25_CORR when all 5 valid")
    add("POST_Knowledge5_Total", "Derived from matched post-session responses", "Post-session total on five repeated knowledge items (Q11-Q14, Q25)", "Numeric", "0-5", "Scale", "Sum of POST_Q11_CORR, POST_Q12_CORR, POST_Q13_CORR, POST_Q14_CORR, POST_Q25_CORR")
    add("Knowledge5_Gain", "Derived from PRE_Knowledge5_Total and POST_Knowledge5_Total", "Matched five-item gain score", "Numeric", "-5 to +5; blank if baseline incomplete", "Scale", "POST_Knowledge5_Total - PRE_Knowledge5_Total")
    add("KnowledgeMatched_Available_NValid", "Derived from baseline-answered matched items", "Number of repeated matched items answered at baseline for that participant", "Numeric", "0-5", "Scale", "Count of nonmissing PRE_Q11_CORR, PRE_Q12_CORR, PRE_Q13_CORR, PRE_Q14_CORR, PRE_Q25_CORR")
    add("PRE_KnowledgeMatched_Available_Total", "Derived from baseline-answered matched items", "Baseline total across the exact repeated items answered at baseline", "Numeric", "0-5; blank if no baseline items answered", "Scale", "Sum of PRE_Q11_CORR, PRE_Q12_CORR, PRE_Q13_CORR, PRE_Q14_CORR, PRE_Q25_CORR over nonmissing baseline items only")
    add("POST_KnowledgeMatched_Available_Total", "Derived from the same matched baseline-answered item subset", "Post-session total across the exact repeated items answered at baseline", "Numeric", "0-5; blank if no baseline items answered or if a corresponding post item is missing", "Scale", "Sum of POST_Q11_CORR, POST_Q12_CORR, POST_Q13_CORR, POST_Q14_CORR, POST_Q25_CORR over the participant's nonmissing baseline subset")
    add("KnowledgeMatched_Available_Gain", "Derived from PRE_KnowledgeMatched_Available_Total and POST_KnowledgeMatched_Available_Total", "Gain across the participant-specific matched baseline-answered subset", "Numeric", "-5 to +5; blank if no matched subset is available", "Scale", "POST_KnowledgeMatched_Available_Total - PRE_KnowledgeMatched_Available_Total")
    add("PRE_KnowledgeMatched_Available_Percent", "Derived from PRE_KnowledgeMatched_Available_Total and KnowledgeMatched_Available_NValid", "Baseline percent across the participant-specific matched available-item subset", "Numeric", "0-100; blank if no baseline items answered", "Scale", "PRE_KnowledgeMatched_Available_Total / KnowledgeMatched_Available_NValid * 100")
    add("POST_KnowledgeMatched_Available_Percent", "Derived from POST_KnowledgeMatched_Available_Total and KnowledgeMatched_Available_NValid", "Post-session percent across the participant-specific matched available-item subset", "Numeric", "0-100; blank if no baseline items answered", "Scale", "POST_KnowledgeMatched_Available_Total / KnowledgeMatched_Available_NValid * 100")
    add("KnowledgeMatched_Available_GainPercent", "Derived from PRE_KnowledgeMatched_Available_Percent and POST_KnowledgeMatched_Available_Percent", "Percent gain across the participant-specific matched available-item subset", "Numeric", "-100 to +100; blank if no matched subset is available", "Scale", "POST_KnowledgeMatched_Available_Percent - PRE_KnowledgeMatched_Available_Percent")
    add("POST_Knowledge15_Total", "Derived from POST_Q11_CORR-POST_Q25_CORR", "Full post-session fifteen-item knowledge total", "Numeric", "0-15", "Scale", "Sum of POST_Q11_CORR:POST_Q25_CORR")
    add("POST_Knowledge15_Percent", "Derived from POST_Knowledge15_Total", "Full post-session knowledge percent", "Numeric", "0-100", "Scale", "POST_Knowledge15_Total / 15 * 100")
    add("POST_KnowledgeConfidence", "Forms Q26", "Post-session confidence in knowledge-check answers", "Numeric", "1=Not at all confident ... 5=Very confident", "Ordinal")

    return metadata


def build_spss_metadata_syntax() -> str:
    labels: list[tuple[str, str]] = [
        ("ParticipantID", "Participant ID"),
        ("BG_Age18Plus", "Age 18 or older"),
        ("BG_PriorArchitecture", "Previously studied computer architecture or computer organisation"),
        ("BG_PriorVR", "Used a VR headset before"),
        ("BG_VRExperience", "VR experience level"),
        ("BG_PriorConfidence", "Pre-session confidence in computer architecture"),
        ("BG_PriorDifficulty", "Perceived difficulty of core architecture topics"),
        ("BG_Studied_InstructionFlow", "Previously studied topic: Instruction Flow"),
        ("BG_Studied_Pipelining", "Previously studied topic: Pipelining"),
        ("BG_Studied_Stack", "Previously studied topic: Stack"),
        ("BG_Studied_MemoryHierarchy", "Previously studied topic: Memory Hierarchy"),
        ("BG_Studied_Assembly", "Previously studied topic: Assembly"),
        ("BG_Studied_None", "Previously studied topic: None of the above"),
    ]

    for field, label in [
        ("USE01", "The VR system was easy to learn"),
        ("USE02", "I could navigate the environment without difficulty"),
        ("USE03", "The controls felt intuitive"),
        ("USE04", "The visual presentation made the concepts easier to follow"),
        ("USE05", "I could focus on the learning task without being distracted by the interface"),
        ("USE06", "I would feel confident using this system again"),
        ("USE07_Discomfort", "I experienced discomfort, motion sickness, or visual fatigue while using the system"),
        ("USE07_Comfort_R", "Reverse-coded comfort item"),
        ("UsabilityMean_6", "Mean of usability items USE01-USE06"),
        ("UsabilityMean_7R", "Mean of usability items USE01-USE06 plus reverse-coded comfort"),
        ("ENG01", "The experience held my attention"),
        ("ENG02", "I felt actively involved in the learning process"),
        ("ENG03", "The system made the topic feel more interesting"),
        ("ENG04", "I was motivated to explore the concepts further"),
        ("ENG05", "The interactive format improved my focus compared with standard learning materials"),
        ("EngagementMean", "Mean of engagement items ENG01-ENG05"),
        ("LEARN01", "The VR experience helped me understand the topic better"),
        ("LEARN02", "The visualisations helped me build a mental model of what was happening"),
        ("LEARN03", "The system made abstract processes easier to understand"),
        ("LEARN04", "I would prefer this as a supplement to traditional teaching materials"),
        ("LEARN05", "I believe I learned something useful from this session"),
        ("LearningExperienceMean", "Mean of learning-experience items LEARN01-LEARN05"),
    ]:
        labels.append((field, label))

    for question_number in BASELINE_QUESTIONS:
        labels.append((f"PRE_Q{question_number}_RESP", f"Baseline response Q{question_number}"))
        labels.append((f"PRE_Q{question_number}_CORR", f"Baseline correctness Q{question_number}"))
    for question in MCQ_QUESTIONS:
        labels.append((f"POST_Q{question.number}_RESP", f"Post-session response Q{question.number}: {question.prompt}"))
        labels.append((f"POST_Q{question.number}_CORR", f"Post-session correctness Q{question.number}"))

    labels.extend(
        [
            ("PRE_Knowledge5_NValid", "Count of nonmissing baseline items Q11, Q12, Q13, Q14, Q25"),
            ("PRE_Knowledge5_Total", "Baseline matched five-item knowledge total"),
            ("POST_Knowledge5_Total", "Post-session total on five repeated knowledge items (Q11-Q14, Q25)"),
            ("Knowledge5_Gain", "Matched five-item knowledge gain"),
            ("KnowledgeMatched_Available_NValid", "Number of repeated matched items answered at baseline"),
            ("PRE_KnowledgeMatched_Available_Total", "Baseline total across the participant-specific matched available-item subset"),
            ("POST_KnowledgeMatched_Available_Total", "Post-session total across the participant-specific matched available-item subset"),
            ("KnowledgeMatched_Available_Gain", "Gain across the participant-specific matched available-item subset"),
            ("PRE_KnowledgeMatched_Available_Percent", "Baseline percent across the participant-specific matched available-item subset"),
            ("POST_KnowledgeMatched_Available_Percent", "Post-session percent across the participant-specific matched available-item subset"),
            ("KnowledgeMatched_Available_GainPercent", "Percent gain across the participant-specific matched available-item subset"),
            ("POST_Knowledge15_Total", "Full post-session knowledge total"),
            ("POST_Knowledge15_Percent", "Full post-session knowledge percent"),
            ("POST_KnowledgeConfidence", "Post-session confidence in knowledge-check answers"),
        ]
    )

    nominal_fields = [
        "ParticipantID",
        "BG_Age18Plus",
        "BG_PriorArchitecture",
        "BG_PriorVR",
        "BG_Studied_InstructionFlow",
        "BG_Studied_Pipelining",
        "BG_Studied_Stack",
        "BG_Studied_MemoryHierarchy",
        "BG_Studied_Assembly",
        "BG_Studied_None",
    ] + [f"PRE_Q{question_number}_RESP" for question_number in BASELINE_QUESTIONS] + [f"PRE_Q{question_number}_CORR" for question_number in BASELINE_QUESTIONS] + [f"POST_Q{question.number}_RESP" for question in MCQ_QUESTIONS] + [f"POST_Q{question.number}_CORR" for question in MCQ_QUESTIONS]

    ordinal_fields = [
        "BG_VRExperience",
        "BG_PriorConfidence",
        "BG_PriorDifficulty",
        "POST_KnowledgeConfidence",
        "USE01",
        "USE02",
        "USE03",
        "USE04",
        "USE05",
        "USE06",
        "USE07_Discomfort",
        "USE07_Comfort_R",
        "ENG01",
        "ENG02",
        "ENG03",
        "ENG04",
        "ENG05",
        "LEARN01",
        "LEARN02",
        "LEARN03",
        "LEARN04",
        "LEARN05",
    ]

    scale_fields = [
        "UsabilityMean_6",
        "UsabilityMean_7R",
        "EngagementMean",
        "LearningExperienceMean",
        "PRE_Knowledge5_NValid",
        "PRE_Knowledge5_Total",
        "POST_Knowledge5_Total",
        "Knowledge5_Gain",
        "KnowledgeMatched_Available_NValid",
        "PRE_KnowledgeMatched_Available_Total",
        "POST_KnowledgeMatched_Available_Total",
        "KnowledgeMatched_Available_Gain",
        "PRE_KnowledgeMatched_Available_Percent",
        "POST_KnowledgeMatched_Available_Percent",
        "KnowledgeMatched_Available_GainPercent",
        "POST_Knowledge15_Total",
        "POST_Knowledge15_Percent",
    ]

    lines: list[str] = [
        "* Auto-generated by build_study_data.py.",
        "VARIABLE LABELS",
    ]
    for variable, label in labels:
        lines.append(f" {variable} {quoted(label)}")
    lines[-1] += "."

    yes_no_targets = ["BG_Age18Plus", "BG_PriorArchitecture", "BG_PriorVR", "BG_Studied_InstructionFlow", "BG_Studied_Pipelining", "BG_Studied_Stack", "BG_Studied_MemoryHierarchy", "BG_Studied_Assembly", "BG_Studied_None"]
    lines.append("")
    lines.append(f"VALUE LABELS {' '.join(yes_no_targets)}")
    lines.append(" 0 'No / Not selected'")
    lines.append(" 1 'Yes / Selected'.")

    lines.append("")
    lines.append("VALUE LABELS BG_VRExperience")
    lines.append(" 1 'None'")
    lines.append(" 2 'Limited'")
    lines.append(" 3 'Moderate'")
    lines.append(" 4 'Good'")
    lines.append(" 5 'High'.")

    lines.append("")
    lines.append("VALUE LABELS BG_PriorConfidence")
    lines.append(" 1 'Extremely not confident'")
    lines.append(" 2 'Somewhat not confident'")
    lines.append(" 3 'Neutral'")
    lines.append(" 4 'Somewhat confident'")
    lines.append(" 5 'Extremely confident'.")

    lines.append("")
    lines.append("VALUE LABELS BG_PriorDifficulty")
    lines.append(" 1 'Extremely difficult'")
    lines.append(" 2 'Somewhat difficult'")
    lines.append(" 3 'Neutral'")
    lines.append(" 4 'Somewhat easy'")
    lines.append(" 5 'Extremely easy'.")

    likert_targets = [
        "USE01", "USE02", "USE03", "USE04", "USE05", "USE06", "USE07_Discomfort",
        "ENG01", "ENG02", "ENG03", "ENG04", "ENG05",
        "LEARN01", "LEARN02", "LEARN03", "LEARN04", "LEARN05",
    ]
    lines.append("")
    lines.append(f"VALUE LABELS {' '.join(likert_targets)}")
    lines.append(" 1 'Strongly Disagree'")
    lines.append(" 2 'Disagree'")
    lines.append(" 3 'Neutral'")
    lines.append(" 4 'Agree'")
    lines.append(" 5 'Strongly Agree'.")

    lines.append("")
    lines.append("VALUE LABELS USE07_Comfort_R")
    lines.append(" 1 'Very uncomfortable'")
    lines.append(" 2 'Uncomfortable'")
    lines.append(" 3 'Neutral'")
    lines.append(" 4 'Comfortable'")
    lines.append(" 5 'Very comfortable'.")

    lines.append("")
    lines.append("VALUE LABELS POST_KnowledgeConfidence")
    lines.append(" 1 'Not at all confident'")
    lines.append(" 2 '2'")
    lines.append(" 3 '3'")
    lines.append(" 4 '4'")
    lines.append(" 5 'Very confident'.")

    for question_number in BASELINE_QUESTIONS:
        question = MCQ_BY_NUMBER[question_number]
        lines.append("")
        lines.append(f"VALUE LABELS PRE_Q{question_number}_RESP")
        for code, option in enumerate(question.options, start=1):
            lines.append(f" {code} {quoted(option)}")
        lines[-1] += "."

    for question in MCQ_QUESTIONS:
        lines.append("")
        lines.append(f"VALUE LABELS POST_Q{question.number}_RESP")
        for code, option in enumerate(question.options, start=1):
            lines.append(f" {code} {quoted(option)}")
        lines[-1] += "."

    lines.append("")
    correctness_fields = " ".join([f"PRE_Q{number}_CORR" for number in BASELINE_QUESTIONS] + [f"POST_Q{question.number}_CORR" for question in MCQ_QUESTIONS])
    lines.append(f"VALUE LABELS {correctness_fields}")
    lines.append(" 0 'Incorrect'")
    lines.append(" 1 'Correct'.")

    lines.append("")
    lines.append("VARIABLE LEVEL")
    for variable in nominal_fields:
        lines.append(f" {variable} (NOMINAL)")
    for variable in ordinal_fields:
        lines.append(f" {variable} (ORDINAL)")
    for variable in scale_fields:
        lines.append(f" {variable} (SCALE)")
    lines[-1] += "."

    decimal_format_fields = {
        "UsabilityMean_6",
        "UsabilityMean_7R",
        "EngagementMean",
        "LearningExperienceMean",
        "PRE_KnowledgeMatched_Available_Percent",
        "POST_KnowledgeMatched_Available_Percent",
        "KnowledgeMatched_Available_GainPercent",
        "POST_Knowledge15_Percent",
    }
    integer_fields = [variable for variable in VARIABLE_ORDER if variable not in decimal_format_fields]
    decimal_fields = [variable for variable in VARIABLE_ORDER if variable in decimal_format_fields]

    lines.append("")
    lines.append(f"FORMATS {' '.join(integer_fields)} (F8.0).")
    lines.append(f"FORMATS {' '.join(decimal_fields)} (F8.2).")
    lines.append("EXECUTE.")

    return "\n".join(lines) + "\n"


def write_csv(path: Path, rows: list[OrderedDict[str, Any]], headers: list[str]) -> None:
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerow(headers)
        for row in rows:
            writer.writerow([csv_scalar(row.get(header)) for header in headers])


def write_markdown_table(path: Path, metadata_rows: list[dict[str, str]]) -> None:
    headers = ["Variable", "Source", "Label", "Type", "ValidValues", "Level", "Derivation"]
    lines = [
        "# Study Data Dictionary",
        "",
        f"Generated: {datetime.now().isoformat(timespec='seconds')}",
        "",
        "| " + " | ".join(headers) + " |",
        "| " + " | ".join(["---"] * len(headers)) + " |",
    ]

    for row in metadata_rows:
        values = [row[header].replace("\n", " ") for header in headers]
        lines.append("| " + " | ".join(values) + " |")

    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def build_quantitative_workbook_rows(quantitative_rows: list[OrderedDict[str, Any]], metadata_rows: list[dict[str, str]]) -> dict[str, Any]:
    data_matrix = [VARIABLE_ORDER] + [[row[column] for column in VARIABLE_ORDER] for row in quantitative_rows]
    dict_headers = ["Variable", "Source", "Label", "Type", "ValidValues", "Level", "Derivation"]
    dictionary_matrix = [dict_headers] + [[row[header] for header in dict_headers] for row in metadata_rows]

    return {
        "output_path": str(QUANT_DIR / "quantitative_master.xlsx"),
        "sheets": [
            {
                "name": "QuantitativeMaster",
                "rows": data_matrix,
                "column_widths": [14] * len(VARIABLE_ORDER),
                "wrap_text": False,
            },
            {
                "name": "DataDictionary",
                "rows": dictionary_matrix,
                "column_widths": [20, 24, 56, 12, 36, 12, 42],
                "wrap_text": True,
            },
        ],
    }


def build_simple_workbook_rows(output_path: Path, headers: list[str], row_dicts: list[dict[str, Any]], column_widths: list[int], sheet_name: str) -> dict[str, Any]:
    matrix = [headers] + [[row.get(header) for header in headers] for row in row_dicts]
    return {
        "output_path": str(output_path),
        "sheets": [
            {
                "name": sheet_name,
                "rows": matrix,
                "column_widths": column_widths,
                "wrap_text": True,
            }
        ],
    }


def build_processing_log(
    source_messages: list[str],
    forms_rows: list[dict[str, Any]],
    baseline_sections: dict[int, dict[str, Any]],
    quantitative_rows: list[OrderedDict[str, Any]],
    validation_audit: dict[str, dict[str, int]],
    baseline_missing_counts: dict[str, int],
    warnings: list[str],
    produced_files: list[Path],
) -> str:
    derived_fields = [
        "USE07_Comfort_R",
        "UsabilityMean_6",
        "UsabilityMean_7R",
        "EngagementMean",
        "LearningExperienceMean",
        "PRE_Knowledge5_NValid",
        "PRE_Knowledge5_Total",
        "POST_Knowledge5_Total",
        "Knowledge5_Gain",
        "KnowledgeMatched_Available_NValid",
        "PRE_KnowledgeMatched_Available_Total",
        "POST_KnowledgeMatched_Available_Total",
        "KnowledgeMatched_Available_Gain",
        "PRE_KnowledgeMatched_Available_Percent",
        "POST_KnowledgeMatched_Available_Percent",
        "KnowledgeMatched_Available_GainPercent",
        "POST_Knowledge15_Total",
        "POST_Knowledge15_Percent",
        "POST_KnowledgeConfidence",
    ]

    lines = [
        "# Processing Log",
        "",
        f"Generated: {datetime.now().isoformat(timespec='seconds')}",
        "",
        "## Source Files",
        f"- Forms export: `{FORMS_PATH}`",
        f"- Baseline notes: `{SESSION_NOTES_PATH}`",
        f"- Questionnaire source: `{QUESTIONNAIRE_PATH}`",
        f"- Master guide: `{MASTER_GUIDE_PATH}`",
        "",
        "## Source Verification",
    ]
    lines.extend([f"- {message}" for message in source_messages])
    lines.extend(
        [
            "",
            "## Validation Summary",
            f"- Forms participants: {len(forms_rows)}",
            f"- Baseline participants: {len(baseline_sections)}",
            f"- IDs successfully merged: {', '.join(str(row['ParticipantID']) for row in quantitative_rows)}",
            f"- Duplicate IDs: 0",
            f"- Unknown response mappings: 0",
            "- Baseline PreScreen mapping: position 1 = Q11, 2 = Q12, 3 = Q13, 4 = Q14, 5 = Q25.",
            "- Baseline probe responses are optional; N/A or blank baseline answers are preserved as missing and are not scored as incorrect.",
            f"- Missing baseline PRE_Q11_RESP: {baseline_missing_counts['PRE_Q11_RESP']}",
            f"- Missing baseline PRE_Q12_RESP: {baseline_missing_counts['PRE_Q12_RESP']}",
            f"- Missing baseline PRE_Q13_RESP: {baseline_missing_counts['PRE_Q13_RESP']}",
            f"- Missing baseline PRE_Q14_RESP: {baseline_missing_counts['PRE_Q14_RESP']}",
            f"- Missing baseline PRE_Q25_RESP: {baseline_missing_counts['PRE_Q25_RESP']}",
            "- Q26 confidence is post-session only and remains separate from all knowledge totals and gain scores.",
            "",
            "## Repeated-Item Matched Counts",
        ]
    )

    for question_number in MATCHED_KNOWLEDGE5_QUESTIONS:
        matched_count = sum(
            1
            for row in quantitative_rows
            if row[f"PRE_Q{question_number}_CORR"] is not None and row[f"POST_Q{question_number}_CORR"] is not None
        )
        lines.append(f"- Q{question_number} matched N: {matched_count}")

    lines.extend(
        [
            "",
            "## Warnings",
        ]
    )

    if warnings:
        lines.extend([f"- {warning}" for warning in warnings])
    else:
        lines.append("- None.")

    lines.extend(["", "## Derived Score Ranges"])
    for field in derived_fields:
        values = [row[field] for row in quantitative_rows if row[field] is not None]
        if not values:
            lines.append(f"- {field}: no valid values")
            continue
        lines.append(f"- {field}: n={len(values)}; min={min(values)}; max={max(values)}")

    lines.extend(["", "## Missingness Audit"])
    lines.append("| Variable | Valid | Missing | Unexpected |")
    lines.append("| --- | ---: | ---: | ---: |")
    for field in VARIABLE_ORDER:
        audit = validation_audit[field]
        lines.append(f"| {field} | {audit['valid']} | {audit['missing']} | {audit['unexpected']} |")

    lines.extend(["", "## Output Files Produced"])
    for produced in produced_files:
        lines.append(f"- `{produced}`")

    return "\n".join(lines) + "\n"


def build_workbook_spec(
    quantitative_rows: list[OrderedDict[str, Any]],
    metadata_rows: list[dict[str, str]],
    questionnaire_feedback_rows: list[dict[str, Any]],
    session_note_rows: list[dict[str, Any]],
) -> dict[str, Any]:
    questionnaire_headers = ["ParticipantID", "QuestionID", "Prompt", "Response", "SourceType"]
    session_headers = ["ParticipantID", "SourceType", "NoteCategory", "Text"]

    return {
        "workbooks": [
            build_quantitative_workbook_rows(quantitative_rows, metadata_rows),
            build_simple_workbook_rows(
                QUAL_DIR / "NVivo_questionnaire_feedback.xlsx",
                questionnaire_headers,
                questionnaire_feedback_rows,
                [14, 12, 56, 100, 22],
                "QuestionnaireFeedback",
            ),
            build_simple_workbook_rows(
                QUAL_DIR / "NVivo_session_notes_anonymized.xlsx",
                session_headers,
                session_note_rows,
                [14, 22, 24, 110],
                "SessionNotes",
            ),
        ]
    }


def run_workbook_builder(spec_path: Path) -> None:
    if not NODE_EXECUTABLE.exists():
        raise ValidationError(f"Node executable for workbook export not found: {NODE_EXECUTABLE}")
    if not WORKBOOK_BUILDER_PATH.exists():
        raise ValidationError(f"Workbook builder script not found: {WORKBOOK_BUILDER_PATH}")

    subprocess.run([str(NODE_EXECUTABLE), str(WORKBOOK_BUILDER_PATH), str(spec_path)], check=True)


def run_likert_figure_builder(input_csv_path: Path) -> list[Path]:
    if not input_csv_path.exists():
        raise ValidationError(f"Quantitative CSV for Likert figure generation not found: {input_csv_path}")
    if not LIKERT_FIGURE_SCRIPT_PATH.exists():
        raise ValidationError(f"Likert figure script not found: {LIKERT_FIGURE_SCRIPT_PATH}")
    if not LIKERT_PLOT_PYTHON:
        raise ValidationError("No Python executable is available for Likert figure generation.")

    FIGURES_DIR.mkdir(parents=True, exist_ok=True)

    try:
        subprocess.run(
            [
                LIKERT_PLOT_PYTHON,
                "-c",
                "import matplotlib, numpy, pandas",
            ],
            check=True,
            capture_output=True,
            text=True,
        )
    except subprocess.CalledProcessError as exc:
        stderr = exc.stderr.strip() if exc.stderr else "unknown import failure"
        raise ValidationError(
            f"Likert plotting runtime '{LIKERT_PLOT_PYTHON}' is missing required packages: {stderr}"
        ) from exc

    subprocess.run(
        [
            LIKERT_PLOT_PYTHON,
            str(LIKERT_FIGURE_SCRIPT_PATH),
            "--input",
            str(input_csv_path),
            "--output-dir",
            str(FIGURES_DIR),
        ],
        check=True,
    )

    participant_count = sum(1 for _ in input_csv_path.open("r", encoding="utf-8")) - 1
    generated_paths: list[Path] = []
    for basename in LIKERT_FIGURE_BASENAMES:
        output_path = FIGURES_DIR / f"{basename}_{participant_count}P.png"
        if not output_path.exists():
            raise ValidationError(f"Expected Likert figure output not found after generation: {output_path}")
        generated_paths.append(output_path)

    return generated_paths


def main() -> int:
    QUANT_DIR.mkdir(parents=True, exist_ok=True)
    QUAL_DIR.mkdir(parents=True, exist_ok=True)
    DOC_DIR.mkdir(parents=True, exist_ok=True)
    TMP_DIR.mkdir(parents=True, exist_ok=True)
    FIGURES_DIR.mkdir(parents=True, exist_ok=True)

    source_messages = verify_source_files()
    forms_rows, warnings = load_forms_rows()
    baseline_sections = parse_session_notes()
    quantitative_rows, questionnaire_feedback_rows, session_note_rows, baseline_missing_counts, baseline_warnings = build_quantitative_rows(forms_rows, baseline_sections)
    warnings.extend(baseline_warnings)
    validation_audit = validate_quantitative_rows(quantitative_rows)
    metadata_rows = build_variable_metadata()

    spss_csv_path = QUANT_DIR / "SPSS_quantitative.csv"
    write_csv(spss_csv_path, quantitative_rows, VARIABLE_ORDER)
    figure_paths = run_likert_figure_builder(spss_csv_path)

    metadata_syntax_path = QUANT_DIR / "SPSS_metadata_setup.sps"
    metadata_syntax_path.write_text(build_spss_metadata_syntax(), encoding="utf-8")

    data_dictionary_path = DOC_DIR / "data_dictionary.md"
    write_markdown_table(data_dictionary_path, metadata_rows)

    workbook_spec_path = TMP_DIR / "workbook_spec.json"
    workbook_spec_path.write_text(json.dumps(build_workbook_spec(quantitative_rows, metadata_rows, questionnaire_feedback_rows, session_note_rows), ensure_ascii=False, indent=2), encoding="utf-8")
    run_workbook_builder(workbook_spec_path)

    produced_files = [
        QUANT_DIR / "quantitative_master.xlsx",
        QUANT_DIR / "SPSS_quantitative.csv",
        QUANT_DIR / "SPSS_metadata_setup.sps",
        QUAL_DIR / "NVivo_questionnaire_feedback.xlsx",
        QUAL_DIR / "NVivo_session_notes_anonymized.xlsx",
        DOC_DIR / "data_dictionary.md",
    ]
    produced_files.extend(figure_paths)

    processing_log_path = QUANT_DIR / "processing_log.md"
    processing_log_path.write_text(
        build_processing_log(
            source_messages,
            forms_rows,
            baseline_sections,
            quantitative_rows,
            validation_audit,
            baseline_missing_counts,
            warnings,
            produced_files,
        ),
        encoding="utf-8",
    )

    print(f"Forms participants: {len(forms_rows)}")
    print(f"Baseline participants: {len(baseline_sections)}")
    print("Merged IDs: " + ", ".join(str(row["ParticipantID"]) for row in quantitative_rows))
    print(f"Missing baseline PRE_Q14_RESP: {baseline_missing_counts['PRE_Q14_RESP']}")
    print(f"Missing baseline PRE_Q25_RESP: {baseline_missing_counts['PRE_Q25_RESP']}")
    print("Validation: PASSED")
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except ValidationError as exc:
        print(f"VALIDATION ERROR: {exc}", file=sys.stderr)
        raise SystemExit(1)
